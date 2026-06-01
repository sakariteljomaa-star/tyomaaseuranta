"""
Tuottaa kustannusseuranta-Excelin Valteri-mallin mukaisella rakenteella.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Värit
OTSIKKO_TAUSTA = "1F3864"   # tumma sininen
OTSIKKO_TEKSTI = "FFFFFF"
ALAOTSIKKO_TAUSTA = "BDD7EE"
SUMMA_TAUSTA = "D9E1F2"
KULULAJI_TAUSTA = "E2EFDA"   # vihreä – kululajiotsikko
HARMAA = "F2F2F2"

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _otsikko(ws, rivi, teksti, sarakkeet=5):
    s = ws.cell(rivi, 1, teksti)
    s.font = Font(bold=True, color=OTSIKKO_TEKSTI, size=12)
    s.fill = PatternFill("solid", fgColor=OTSIKKO_TAUSTA)
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=rivi, start_column=1, end_row=rivi, end_column=sarakkeet)
    ws.row_dimensions[rivi].height = 22


def _kenttaotsikot(ws, rivi, otsikot):
    for j, ot in enumerate(otsikot, 1):
        c = ws.cell(rivi, j, ot)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[rivi].height = 30


def _summarivi(ws, rivi, label, arvo, sarakkeet=5):
    c = ws.cell(rivi, 1, label)
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor=SUMMA_TAUSTA)
    v = ws.cell(rivi, sarakkeet, arvo)
    v.font = Font(bold=True)
    v.fill = PatternFill("solid", fgColor=SUMMA_TAUSTA)
    v.number_format = '#,##0.00 "€"'


def _datarivi(ws, rivi, arvot, tausta=None):
    for j, a in enumerate(arvot, 1):
        c = ws.cell(rivi, j, a)
        c.border = THIN_BORDER
        if tausta:
            c.fill = PatternFill("solid", fgColor=tausta)
        if isinstance(a, (int, float)) and j > 1:
            c.number_format = '#,##0.00'
            c.alignment = Alignment(horizontal="right")


def _kululaji_otsikko(ws, rivi, teksti, sarakkeet=5):
    ws.merge_cells(start_row=rivi, start_column=1, end_row=rivi, end_column=sarakkeet)
    c = ws.cell(rivi, 1, f"── {teksti} ──")
    c.font = Font(bold=True, italic=True)
    c.fill = PatternFill("solid", fgColor=KULULAJI_TAUSTA)
    c.alignment = Alignment(horizontal="left")


def luo_excel(
    df: pd.DataFrame,
    projekti: str = "",
    yritys: str = "Uudenmaan Asbestipurku Oy",
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    _tee_ostot(wb, df, projekti, yritys)
    _tee_jatemaksut(wb, df, projekti, yritys)
    _tee_aliurakoitsijat(wb, df, projekti, yritys)
    _tee_yhteenveto(wb, df, projekti, yritys)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── OSTOT ─────────────────────────────────────────────────────────────────────

def _tee_ostot(wb, df, projekti, yritys):
    ws = wb.create_sheet("Ostot")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 16

    r = 1
    _otsikko(ws, r, f"OSTOSEURANTA – {projekti}", 7); r += 1
    ws.cell(r, 1, f"{yritys}  |  Lähde: Netvisor ostoreskontra  |  Hinnat alv 0%")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7); r += 1
    ws.cell(r, 1, "ALV-kohtelu: KOOS = normaali 25,5% (vähennetään ALV-tilityksessä)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7); r += 1
    r += 1

    _kenttaotsikot(ws, r, ["Päivä", "Tosite", "Lasku", "Nimike / toimittaja",
                            "Summa\n(alv 0%)", "Kategoria", "Nimike-laji"]); r += 1

    ostot = df[df["kululaji"] == "Ostot"].copy()

    for kat in ["Lisätyö", "Urakka"]:
        osa = ostot[ostot["kategoria"] == kat]
        if osa.empty:
            continue
        _kululaji_otsikko(ws, r, kat.upper(), 7); r += 1
        for _, row in osa.iterrows():
            _datarivi(ws, r, [
                row["pvm"].strftime("%d.%m") if pd.notna(row["pvm"]) else "",
                row["tosite"],
                row["lasku"],
                row["selite"],
                row["summa"],
                row["kategoria"],
                row["nimike_laji"],
            ])
            r += 1
        summa = osa["summa"].sum()
        _summarivi(ws, r, f"{kat} yhteensä (alv 0%)", summa, 5); r += 2

    _summarivi(ws, r, "OSTOT YHTEENSÄ (alv 0%)", ostot["summa"].sum(), 5)


# ── JÄTEMAKSUT ────────────────────────────────────────────────────────────────

def _tee_jatemaksut(wb, df, projekti, yritys):
    ws = wb.create_sheet("Jätemaksut")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12

    r = 1
    _otsikko(ws, r, f"JÄTEMAKSUT – {projekti}", 6); r += 1
    ws.cell(r, 1, f"{yritys}  |  Kuljetusrinki Oy = jätemaksut  |  ALV KOOS 25,5%  |  Hinnat alv 0%")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6); r += 2

    _kenttaotsikot(ws, r, ["Päivä", "Tosite", "Lasku", "Toimittaja / selite",
                            "Summa\n(alv 0%)", "Kategoria"]); r += 1

    jate = df[df["kululaji"] == "Jätemaksut"].copy()

    for kat in ["Lisätyö", "Urakka"]:
        osa = jate[jate["kategoria"] == kat]
        if osa.empty:
            continue
        _kululaji_otsikko(ws, r, kat.upper(), 6); r += 1
        for _, row in osa.iterrows():
            _datarivi(ws, r, [
                row["pvm"].strftime("%d.%m") if pd.notna(row["pvm"]) else "",
                row["tosite"],
                row["lasku"],
                row["selite"],
                row["summa"],
                row["kategoria"],
            ])
            r += 1
        summa = osa["summa"].sum()
        _summarivi(ws, r, f"{kat} jätemaksut yht.", summa, 5); r += 2

    _summarivi(ws, r, "JÄTEMAKSUT YHTEENSÄ (alv 0%)", jate["summa"].sum(), 5)


# ── ALIURAKOITSIJAT ───────────────────────────────────────────────────────────

def _tee_aliurakoitsijat(wb, df, projekti, yritys):
    ws = wb.create_sheet("Aliurakoitsijat")
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 30

    r = 1
    _otsikko(ws, r, f"ALIURAKOITSIJAT – {projekti}", 8); r += 1
    ws.cell(r, 1, f"{yritys}  |  RAOS = käänteinen rakennusalan ALV  |  Ei sivukulukerrointa")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8); r += 2

    _kenttaotsikot(ws, r, ["Päivä", "Tosite", "Lasku", "Toimittaja / selite",
                            "Summa\n(alv 0%)", "Huomio", "Kategoria", "Laskentakohde"]); r += 1

    ali = df[df["kululaji"] == "Aliurakoitsijat"].copy()

    for kat in ["Lisätyö", "Urakka"]:
        osa = ali[ali["kategoria"] == kat]
        if osa.empty:
            continue
        _kululaji_otsikko(ws, r, kat.upper(), 8); r += 1
        for _, row in osa.iterrows():
            alv_huomio = "RAOS" if str(row.get("alv_tunnus", "")).upper() == "RAOS" else "ALV 0%"
            _datarivi(ws, r, [
                row["pvm"].strftime("%d.%m") if pd.notna(row["pvm"]) else "",
                row["tosite"],
                row["lasku"],
                row["selite"],
                row["summa"],
                alv_huomio,
                row["kategoria"],
                row["laskentakohteet"] if pd.notna(row.get("laskentakohteet")) else "",
            ])
            r += 1
        summa = osa["summa"].sum()
        _summarivi(ws, r, f"{kat} aliurakointi yht.", summa, 5); r += 2

    _summarivi(ws, r, "ALIURAKOINTI YHTEENSÄ (alv 0%)", ali["summa"].sum(), 5)


# ── YHTEENVETO ────────────────────────────────────────────────────────────────

def _tee_yhteenveto(wb, df, projekti, yritys):
    ws = wb.create_sheet("Yhteenveto")
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14

    r = 1
    _otsikko(ws, r, "KUSTANNUSSEURANTA – YHTEENVETO", 3); r += 1
    ws.cell(r, 1, f"{projekti}  |  {yritys}")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 2

    def rivi(label, arvo, bold=False, tausta=None):
        nonlocal r
        c = ws.cell(r, 1, label)
        v = ws.cell(r, 3, arvo)
        if bold:
            c.font = Font(bold=True)
            v.font = Font(bold=True)
        if tausta:
            c.fill = PatternFill("solid", fgColor=tausta)
            v.fill = PatternFill("solid", fgColor=tausta)
        v.number_format = '#,##0.00 "€"'
        v.alignment = Alignment(horizontal="right")
        r += 1

    def otsikko_rivi(teksti):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        c = ws.cell(r, 1, teksti)
        c.font = Font(bold=True, color=OTSIKKO_TEKSTI)
        c.fill = PatternFill("solid", fgColor="2E75B6")
        r += 1

    ostot = df[df["kululaji"] == "Ostot"]["summa"].sum()
    jate = df[df["kululaji"] == "Jätemaksut"]["summa"].sum()
    kalusto = df[df["kululaji"] == "Kalusto"]["summa"].sum()
    ali = df[df["kululaji"] == "Aliurakoitsijat"]["summa"].sum()
    kaikki = df["summa"].sum()

    kat_urakka = df[df["kategoria"] == "Urakka"]["summa"].sum()
    kat_lisatyo = df[df["kategoria"] == "Lisätyö"]["summa"].sum()

    otsikko_rivi("KUSTANNUKSET KULULAJEITTAIN")
    rivi("Materiaalit / tarvikkeet (KOOS)", ostot)
    rivi("Jätemaksut (KOOS)", jate)
    rivi("Kalusto / konevuokra (KOOS)", kalusto)
    rivi("Aliurakoitsijat (RAOS)", ali)
    r += 1
    rivi("KAIKKI OSTOT YHTEENSÄ (alv 0%)", kaikki, bold=True, tausta=SUMMA_TAUSTA)
    r += 1

    otsikko_rivi("KUSTANNUKSET KOHTEITTAIN")
    rivi("Urakka-kustannukset (alv 0%)", kat_urakka)
    rivi("Lisätyö-kustannukset (alv 0%)", kat_lisatyo)
    r += 1
    rivi("YHTEENSÄ (alv 0%)", kaikki, bold=True, tausta=SUMMA_TAUSTA)
    r += 1

    otsikko_rivi("ALV-KOHTELU")
    koos_summa = df[df["alv_tunnus"].str.upper() == "KOOS"]["summa"].sum() if "alv_tunnus" in df.columns else 0
    raos_summa = df[df["alv_tunnus"].str.upper() == "RAOS"]["summa"].sum() if "alv_tunnus" in df.columns else 0
    rivi("KOOS-ostot (ALV 25,5% vähennetään)", koos_summa)
    rivi("RAOS-ostot (käänteinen ALV, ei kassavaikutusta)", raos_summa)
    rivi("KOOS-ALV-vähennysoikeus (× 25,5%)", koos_summa * 0.255, tausta=KULULAJI_TAUSTA)
