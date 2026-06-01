"""
Aliurakoitsijoiden viikkoraportti — ladattava Excel, tulostettavissa PDF:ksi.
"""

import io
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PAIVAT = ["Ma", "Ti", "Ke", "To", "Pe", "La", "Su"]

THIN = Side(style="thin", color="AAAAAA")
MED  = Side(style="medium", color="444444")
THIN_B = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_B  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

SININEN   = "1F3864"
VAALEASIN = "BDD7EE"
VIHREA    = "E2EFDA"
HARMAA    = "F2F2F2"
KELTAINEN = "FFF2CC"


def _viikon_pvm(vuosi: int, viikko: int):
    """Palauttaa (ma_pvm, su_pvm) annetulle viikolle."""
    ma = date.fromisocalendar(vuosi, viikko, 1)
    return ma, ma + timedelta(days=6)


def _otsikkosolu(ws, rivi, sarake, teksti, bold=True, tausta=None, koko=11, wrap=False, tasaus="center"):
    c = ws.cell(rivi, sarake, teksti)
    c.font = Font(bold=bold, size=koko,
                  color="FFFFFF" if tausta and tausta not in (HARMAA, VAALEASIN, VIHREA, KELTAINEN) else "000000")
    if tausta:
        c.fill = PatternFill("solid", fgColor=tausta)
    c.alignment = Alignment(horizontal=tasaus, vertical="center", wrap_text=wrap)
    c.border = THIN_B
    return c


def _datasolu(ws, rivi, sarake, arvo, numero=False, tausta=None, bold=False):
    c = ws.cell(rivi, sarake, arvo)
    c.font = Font(bold=bold, size=10)
    c.border = THIN_B
    c.alignment = Alignment(horizontal="right" if numero else "left", vertical="center")
    if tausta:
        c.fill = PatternFill("solid", fgColor=tausta)
    if numero and isinstance(arvo, (int, float)):
        c.number_format = '#,##0.00'
    return c


def luo_viikkoraportti(
    rivit: list,
    viikko: int,
    vuosi: int,
    projekti: str = "",
    yritys: str = "Uudenmaan Asbestipurku Oy",
) -> bytes:
    """
    rivit: lista dict-rivejä (kaikki ali_tunnit), suodatetaan viikolle.
    """
    viikkorivit = [r for r in rivit if r.get("viikko") == viikko]

    wb = Workbook()
    ws = wb.active
    ws.title = f"Viikko {viikko}"
    ws.sheet_view.showGridLines = False

    # Sarakeleveydet
    ws.column_dimensions["A"].width = 22   # Nimi
    ws.column_dimensions["B"].width = 20   # Yritys
    for col in ["C","D","E","F","G","H","I"]:  # Ma–Su + Yht
        ws.column_dimensions[col].width = 6
    ws.column_dimensions["J"].width = 14   # Kategoria
    ws.column_dimensions["K"].width = 12   # Laskutustapa
    ws.column_dimensions["L"].width = 11   # €/h tai kiinteä
    ws.column_dimensions["M"].width = 12   # Summa
    ws.column_dimensions["N"].width = 22   # Huomio

    ma_pvm, su_pvm = _viikon_pvm(vuosi, viikko)
    pvm_str = f"{ma_pvm.strftime('%-d.%-m.')} – {su_pvm.strftime('%-d.%-m.%Y')}"

    r = 1

    # ── Otsikkoalue ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:N{r}")
    c = ws.cell(r, 1, f"ALIURAKOITSIJOIDEN TUNTIKIRJA – {projekti.upper()}")
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=SININEN)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 26
    r += 1

    ws.merge_cells(f"A{r}:N{r}")
    c = ws.cell(r, 1, f"{yritys}  |  Viikko {viikko}  |  {pvm_str}")
    c.font = Font(size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=SININEN)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18
    r += 1

    r += 1  # tyhjä rivi

    # ── Sarakeotsikit ─────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 36
    otsikot = ["Nimi", "Yritys"] + PAIVAT + ["Yht\n(h)", "Kategoria", "Laskutus", "Hinta\n(€/h tai €)", "Summa\n(€)", "Tila", "Huomio"]
    for j, ot in enumerate(otsikot, 1):
        _otsikkosolu(ws, r, j, ot, tausta=SININEN, wrap=True)
    r += 1

    # ── Päivämäärät otsikkojen alle ───────────────────────────────────────────
    ws.row_dimensions[r].height = 14
    for d, paiva_idx in enumerate(range(7), 3):
        pvm = ma_pvm + timedelta(days=d - 3)
        c = ws.cell(r, d, pvm.strftime("%-d.%-m"))
        c.font = Font(size=8, color="666666")
        c.alignment = Alignment(horizontal="center")
    r += 1

    if not viikkorivit:
        ws.merge_cells(f"A{r}:N{r}")
        ws.cell(r, 1, "– Ei kirjauksia tälle viikolle –").font = Font(italic=True, color="888888")
    else:
        # Ryhmittele yrityksittäin
        yritykset = {}
        for rv in viikkorivit:
            yr = rv.get("yritys", "–")
            yritykset.setdefault(yr, []).append(rv)

        for yr_nimi, yr_rivit in yritykset.items():
            # Yritys-otsikkorivi
            ws.merge_cells(f"A{r}:N{r}")
            c = ws.cell(r, 1, f"  {yr_nimi}")
            c.font = Font(bold=True, size=10)
            c.fill = PatternFill("solid", fgColor=VAALEASIN)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[r].height = 18
            r += 1

            yr_yht_h = 0.0
            yr_yht_eur = 0.0

            for rv in yr_rivit:
                ws.row_dimensions[r].height = 18
                tausta = None

                _datasolu(ws, r, 1, rv.get("nimi", ""), tausta=tausta)
                _datasolu(ws, r, 2, rv.get("yritys", ""), tausta=tausta)

                yht_h = 0.0
                for d_i, paiva in enumerate(["ma","ti","ke","to","pe","la","su"], 3):
                    h = rv.get(paiva, 0) or 0
                    c = ws.cell(r, d_i, h if h else "")
                    c.border = THIN_B
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.font = Font(size=10)
                    if h:
                        yht_h += h

                # Yht (h)
                c_yht = ws.cell(r, 10, yht_h)
                c_yht.font = Font(bold=True, size=10)
                c_yht.border = THIN_B
                c_yht.alignment = Alignment(horizontal="center")
                c_yht.fill = PatternFill("solid", fgColor=HARMAA)

                _datasolu(ws, r, 11, rv.get("kategoria", ""), tausta=tausta)

                laskutustapa = rv.get("laskutustapa", "tunnit")
                if laskutustapa == "tuntihinta":
                    lask_label = "€/h"
                    hinta = rv.get("tuntihinta") or 0
                    summa = yht_h * hinta
                elif laskutustapa == "kiintea":
                    lask_label = "Kiinteä"
                    hinta = rv.get("kiintea_hinta") or 0
                    summa = hinta
                else:
                    lask_label = "Vain h"
                    hinta = None
                    summa = None

                _datasolu(ws, r, 12, lask_label)
                _datasolu(ws, r, 13, hinta if hinta else "", numero=bool(hinta))
                _datasolu(ws, r, 14, summa if summa is not None else "", numero=bool(summa),
                          tausta=VIHREA if summa else None)

                # Hyväksyntätila värillä
                hyv_tila = rv.get("hyvaksynta_tila", "odottaa")
                hyv_label = {"hyvaksytty": "✅ Hyväksytty", "selvitys": "⚠️ Selvitys", "odottaa": "🔵 Odottaa"}.get(hyv_tila, "🔵 Odottaa")
                hyv_väri  = {"hyvaksytty": "E8F5E9", "selvitys": "FFF8E1", "odottaa": "E3F0FF"}.get(hyv_tila, "E3F0FF")
                _datasolu(ws, r, 15, hyv_label, tausta=hyv_väri)

                _datasolu(ws, r, 16, rv.get("huomio", ""))

                yr_yht_h += yht_h
                if summa:
                    yr_yht_eur += summa
                r += 1

                # Päiväkohtaiset huomiot omalle riville (jos on merkintöjä)
                huomiot = rv.get("huomiot", {})
                paiva_avaimet = ["ma","ti","ke","to","pe","la","su"]
                merkinnät = [(i, huomiot.get(pk,"")) for i, pk in enumerate(paiva_avaimet)
                             if huomiot.get(pk,"").strip()]
                if merkinnät:
                    ws.row_dimensions[r].height = 13
                    ws.merge_cells(f"A{r}:B{r}")
                    c = ws.cell(r, 1, "")
                    c.fill = PatternFill("solid", fgColor="F8F8F8")
                    for col_i, teksti in merkinnät:
                        c = ws.cell(r, col_i + 3, teksti)
                        c.font = Font(size=8, italic=True, color="555555")
                        c.alignment = Alignment(horizontal="center", wrap_text=True)
                        c.fill = PatternFill("solid", fgColor="F8F8F8")
                        c.border = THIN_B
                    # Täytä tyhjät solut samalla taustavärillä
                    for col_i in range(3, 16):
                        if ws.cell(r, col_i).value is None:
                            ws.cell(r, col_i).fill = PatternFill("solid", fgColor="F8F8F8")
                            ws.cell(r, col_i).border = THIN_B
                    r += 1

            # Yrityksen yhteensä-rivi
            ws.row_dimensions[r].height = 16
            ws.merge_cells(f"A{r}:I{r}")
            c = ws.cell(r, 1, f"  {yr_nimi} yhteensä")
            c.font = Font(bold=True, size=10)
            c.fill = PatternFill("solid", fgColor=VAALEASIN)
            c.alignment = Alignment(horizontal="left")

            c_yh = ws.cell(r, 10, yr_yht_h)
            c_yh.font = Font(bold=True, size=10)
            c_yh.fill = PatternFill("solid", fgColor=VAALEASIN)
            c_yh.alignment = Alignment(horizontal="center")
            c_yh.border = THIN_B

            ws.merge_cells(f"K{r}:M{r}")
            if yr_yht_eur:
                c_eur = ws.cell(r, 11, f"Summa: {yr_yht_eur:,.2f} €")
                c_eur.font = Font(bold=True, size=10)
                c_eur.fill = PatternFill("solid", fgColor=VAALEASIN)
                c_eur.alignment = Alignment(horizontal="right")
            r += 2

    # ── Yhteenveto ────────────────────────────────────────────────────────────
    r += 1
    ws.merge_cells(f"A{r}:N{r}")
    c = ws.cell(r, 1, "YHTEENVETO")
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=SININEN)
    c.alignment = Alignment(horizontal="left")
    ws.row_dimensions[r].height = 20
    r += 1

    yht_per_kat: dict[str, float] = {}
    yht_per_yr: dict[str, float] = {}
    grand_h = 0.0
    grand_eur = 0.0
    for rv in viikkorivit:
        h = sum(rv.get(p, 0) or 0 for p in ["ma","ti","ke","to","pe","la","su"])
        kat = rv.get("kategoria", "–")
        yr = rv.get("yritys", "–")
        yht_per_kat[kat] = yht_per_kat.get(kat, 0) + h
        yht_per_yr[yr] = yht_per_yr.get(yr, 0) + h
        grand_h += h
        lp = rv.get("laskutustapa", "tunnit")
        if lp == "tuntihinta":
            grand_eur += h * (rv.get("tuntihinta") or 0)
        elif lp == "kiintea":
            grand_eur += rv.get("kiintea_hinta") or 0

    for kat, h in sorted(yht_per_kat.items()):
        ws.cell(r, 1, f"  {kat}").font = Font(size=10)
        ws.cell(r, 10, f"{h:.1f} h").font = Font(bold=True, size=10)
        r += 1

    ws.merge_cells(f"A{r}:I{r}")
    c = ws.cell(r, 1, "  KAIKKI YHTEENSÄ")
    c.font = Font(bold=True, size=11)
    c.fill = PatternFill("solid", fgColor=KELTAINEN)
    c = ws.cell(r, 10, f"{grand_h:.1f} h")
    c.font = Font(bold=True, size=11)
    c.fill = PatternFill("solid", fgColor=KELTAINEN)
    c.alignment = Alignment(horizontal="center")
    if grand_eur:
        c2 = ws.cell(r, 14, grand_eur)
        c2.font = Font(bold=True, size=11)
        c2.fill = PatternFill("solid", fgColor=KELTAINEN)
        c2.number_format = '#,##0.00 "€"'
        c2.alignment = Alignment(horizontal="right")
    ws.row_dimensions[r].height = 22
    r += 3

    # ── Allekirjoitusrivi ─────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(r, 1, "Päiväys / Allekirjoitus").font = Font(size=9, color="888888")
    ws.merge_cells(f"H{r}:N{r}")
    ws.cell(r, 8, "Hyväksytty / Tilaaja").font = Font(size=9, color="888888")
    r += 1
    for col in [1, 8]:
        c = ws.cell(r, col, "_" * 40)
        c.font = Font(size=9, color="AAAAAA")
    r += 1
    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(r, 1, f"{yritys}").font = Font(size=8, color="AAAAAA")

    # Tulostusalue
    ws.print_area = f"A1:N{r}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
